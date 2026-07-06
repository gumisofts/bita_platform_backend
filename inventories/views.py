import csv
import io
import logging

import openpyxl
from django.contrib.postgres.search import TrigramSimilarity
from django.db import transaction
from django.db.models import (
    Count,
    F,
    IntegerField,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Sum,
)
from django.db.models.functions import Coalesce, Lower
from django.http import HttpResponse
from django.utils import timezone
from drf_spectacular.utils import OpenApiParameter, OpenApiTypes, extend_schema
from guardian.shortcuts import assign_perm, get_objects_for_user, get_perms, remove_perm
from rest_framework import status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.exceptions import ValidationError
from rest_framework.mixins import (
    CreateModelMixin,
    DestroyModelMixin,
    ListModelMixin,
    RetrieveModelMixin,
    UpdateModelMixin,
)
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import GenericViewSet, ModelViewSet

from business.models import Branch, Business, biz_perm
from business.permissions import (
    BranchLevelPermission,
    BusinessLevelPermission,
    accessible_branches,
    filter_queryset_by_branch,
)
from core.idempotency import idempotent
from core.utils import is_valid_uuid

from .filters import (
    GroupFilter,
    ItemFilter,
    ItemVariantFilter,
    SuppliedItemFilter,
    SupplierFilter,
    SupplyFilter,
)
from .models import *
from .serializers import *


class ItemViewset(ModelViewSet):
    queryset = Item.objects.all()
    serializer_class = ItemSerializer
    permission_classes = [IsAuthenticated, BranchLevelPermission]
    filterset_class = ItemFilter

    def get_serializer_class(self):
        if self.action == "retrieve":
            return ItemReadSerializer
        return self.serializer_class

    def get_queryset(self):
        total_qty_subquery = (
            SuppliedItem.objects.filter(variant__item=OuterRef("pk"))
            .values("variant__item")
            .annotate(total=Sum("quantity"))
            .values("total")
        )
        queryset = (
            super()
            .get_queryset()
            .select_related("group", "business", "branch")
            .prefetch_related("categories")
            .annotate(
                total_quantity=Coalesce(
                    Subquery(total_qty_subquery, output_field=IntegerField()), 0
                )
            )
        )
        return filter_queryset_by_branch(queryset, self.request, "item")

    # ------------------------------------------------------------------
    # Import / export shared helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_int(value, default=0):
        try:
            return int(str(value).strip())
        except (ValueError, TypeError):
            return default

    @staticmethod
    def _parse_decimal(value):
        try:
            v = float(str(value).strip())
            return v if v > 0 else None
        except (ValueError, TypeError):
            return None

    @staticmethod
    def _rows_from_file(uploaded_file):
        """Parse CSV or Excel upload into a list of dicts keyed by column name."""
        fname = uploaded_file.name.lower()
        if fname.endswith(".csv"):
            text = uploaded_file.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            return [{k.strip(): (v or "") for k, v in row.items()} for row in reader]
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        raw = list(ws.iter_rows(values_only=True))
        if not raw:
            return []
        headers = [str(h).strip() if h is not None else "" for h in raw[0]]
        return [
            {headers[i]: ("" if v is None else str(v)) for i, v in enumerate(row)}
            for row in raw[1:]
        ]

    @staticmethod
    def _build_xlsx_response(wb, filename):
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    @staticmethod
    def _styled_wb_sheet(wb, title):
        """Return (worksheet, header_fill, header_font, thin_border) helpers."""
        ws = wb.active if wb.sheetnames == ["Sheet"] else wb.create_sheet(title)
        ws.title = title
        fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
        font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
        border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin"),
            right=openpyxl.styles.Side(style="thin"),
            top=openpyxl.styles.Side(style="thin"),
            bottom=openpyxl.styles.Side(style="thin"),
        )
        return ws, fill, font, border

    @staticmethod
    def _write_xlsx_headers(ws, columns, fill, font, border):
        center = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = fill
            cell.font = font
            cell.alignment = center
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 24
        ws.row_dimensions[1].height = 25

    # ------------------------------------------------------------------
    # Template download
    # ------------------------------------------------------------------

    @action(
        detail=False,
        methods=["get"],
        url_path="bulk-import-template",
        permission_classes=[IsAuthenticated],
    )
    def bulk_import_template(self, request):
        """Download a styled Excel template for bulk product import."""
        wb = openpyxl.Workbook()
        ws, fill, font, border = self._styled_wb_sheet(wb, "Products")
        self._write_xlsx_headers(ws, BULK_IMPORT_COLUMNS, fill, font, border)

        # Sample rows — two variants for the same product show multi-variant usage
        sample_rows = [
            # name               description           unit    variant_name  price   sku        qty
            [
                "Paracetamol",
                "Pain relief tablet",
                "piece",
                "100mg",
                25.00,
                "SKU-P100",
                "BATCH-001",
                "2024-12-31",
                100,
            ],
            [
                "Paracetamol",
                "",
                "piece",
                "500mg",
                45.00,
                "SKU-P500",
                "BATCH-002",
                "2024-12-31",
                50,
            ],
            [
                "Vitamin C",
                "Immune support",
                "piece",
                "",
                35.00,
                "",
                "BATCH-003",
                "2024-12-31",
                200,
            ],
            [
                "Amoxicillin 250mg",
                "Antibiotic capsule",
                "piece",
                "",
                60.00,
                "SKU-A250",
                "BATCH-004",
                "2024-12-31",
                80,
            ],
        ]
        row_font = openpyxl.styles.Font(size=10)
        for row_idx, row in enumerate(sample_rows, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = row_font
                cell.border = border

        # Instructions sheet
        ws2 = wb.create_sheet("Instructions")
        ws2.column_dimensions["A"].width = 20
        ws2.column_dimensions["B"].width = 70
        title_cell = ws2["A1"]
        title_cell.value = "Bulk Product Import — Column Reference"
        title_cell.font = openpyxl.styles.Font(bold=True, size=13, color="1F4E79")
        ws2.merge_cells("A1:B1")
        ws2.row_dimensions[1].height = 22

        ws2.cell(row=2, column=1, value="Column").font = openpyxl.styles.Font(bold=True)
        ws2.cell(row=2, column=2, value="Description").font = openpyxl.styles.Font(
            bold=True
        )
        for row_idx, col_name in enumerate(BULK_IMPORT_COLUMNS, start=3):
            ws2.cell(row=row_idx, column=1, value=col_name).font = openpyxl.styles.Font(
                bold=True
            )
            ws2.cell(
                row=row_idx, column=2, value=BULK_IMPORT_COLUMN_NOTES.get(col_name, "")
            )

        note_row = 3 + len(BULK_IMPORT_COLUMNS) + 1
        ws2.cell(row=note_row, column=1, value="Multi-variant tip:").font = (
            openpyxl.styles.Font(bold=True, color="C00000")
        )
        ws2.cell(
            row=note_row,
            column=2,
            value="To add multiple variants to one product, repeat the product name on consecutive rows with different variant_name and selling_price values.",
        )

        return self._build_xlsx_response(wb, "bulk_product_import_template.xlsx")

    # ------------------------------------------------------------------
    # Bulk import
    # ------------------------------------------------------------------

    @action(
        detail=False,
        methods=["post"],
        url_path="bulk-import",
        permission_classes=[IsAuthenticated, BranchLevelPermission],
    )
    def bulk_import(self, request):
        """
        Upload a CSV or Excel file to bulk-create products.

        Format rules:
        - One row per variant.
        - Rows sharing the same `name` belong to the same product — the product is
          created from the first row; subsequent rows add more variants.
        - `variant_name` defaults to the product name when left blank (single-variant products).
        - Existing products/variants are upserted (matched by name+branch for items,
          SKU for variants if provided, else by variant name+item).

        Required columns : name, inventory_unit, selling_price
        Optional columns : description, variant_name, sku, quantity
        """
        serializer = BulkItemImportSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        uploaded_file = serializer.validated_data["file"]
        try:
            rows = self._rows_from_file(uploaded_file)
        except Exception as exc:
            return Response(
                {"detail": f"Could not parse file: {exc}"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if not rows:
            return Response(
                {"detail": "The uploaded file contains no data rows."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        business = request.business
        branch = request.branch

        # ------------------------------------------------------------------
        # Group rows by product name preserving order
        # ------------------------------------------------------------------
        from collections import OrderedDict

        groups = OrderedDict()  # name -> list of (row_num, row_dict)
        for row_num, row in enumerate(rows, start=2):
            if not any(str(v).strip() for v in row.values()):
                continue  # skip blank rows
            name = str(row.get("name", "")).strip()
            if not name:
                continue
            groups.setdefault(name, []).append((row_num, row))

        # Stock from the import is recorded as a single supply batch so the
        # quantities are tracked like any other received inventory (and the
        # SuppliedItem post_save signal drives variant.quantity & supply totals).
        # The label is the import moment — identifiable and unique per branch.
        # No payment_method is set, so no purchase/debt transaction is created.
        import_supply_label = f"import-{timezone.now():%Y%m%d-%H%M%S}"
        needs_supply = any(
            self._parse_int(r.get("quantity"), default=0) > 0
            and self._parse_decimal(str(r.get("selling_price", "")).strip()) is not None
            for r in rows
        )
        import_supply = None
        if needs_supply:
            import_supply, _ = Supply.objects.get_or_create(
                branch=branch,
                label=import_supply_label,
                defaults={"business": business},
            )

        created_items = []
        created_variants = []
        errors = []

        for product_name, group_rows in groups.items():
            first_row_num, first_row = group_rows[0]
            inventory_unit = str(first_row.get("inventory_unit", "")).strip()

            if not inventory_unit:
                errors.append(
                    {
                        "row": first_row_num,
                        "name": product_name,
                        "errors": ["'inventory_unit' is required."],
                    }
                )
                continue

            # Resolve groups from the first row (comma-separated names)
            raw_groups = str(first_row.get("groups", "")).strip()
            group_names = [g.strip() for g in raw_groups.split(",") if g.strip()]
            resolved_groups = []
            for gname in group_names:
                grp, _ = Group.objects.get_or_create(name=gname, business=business)
                resolved_groups.append(grp)

            # Upsert Item by name + branch
            try:
                with transaction.atomic():
                    item, item_created = Item.objects.get_or_create(
                        name=product_name,
                        branch=branch,
                        defaults={
                            "description": str(first_row.get("description", "")).strip()
                            or None,
                            "inventory_unit": inventory_unit,
                            "business": business,
                            "group": resolved_groups[0] if resolved_groups else None,
                        },
                    )
                    if not item_created and resolved_groups:
                        item.group = resolved_groups[0]
                        item.save(update_fields=["group"])
                    if item_created:
                        created_items.append(
                            {"row": first_row_num, "name": product_name}
                        )

                    is_first_variant = not item.variants.exists()

                    for row_num, row in group_rows:
                        selling_price_raw = str(row.get("selling_price", "")).strip()
                        selling_price = self._parse_decimal(selling_price_raw)
                        # A blank price is allowed — the variant is still created
                        # with no price (the model permits a null selling_price).
                        # Only a non-blank, unparseable value is an error. This
                        # keeps priceless variants from vanishing on an
                        # export -> import round-trip.
                        if selling_price_raw and selling_price is None:
                            errors.append(
                                {
                                    "row": row_num,
                                    "name": product_name,
                                    "errors": [
                                        "'selling_price' must be a positive number."
                                    ],
                                }
                            )
                            continue

                        variant_name = (
                            str(row.get("variant_name", "")).strip() or product_name
                        )
                        sku = str(row.get("sku", "")).strip() or None
                        quantity = self._parse_int(row.get("quantity"), default=0)

                        batch_number = (
                            str(row.get("batch_number", "")).strip()
                            or import_supply_label
                        )

                        expired_date_str = str(row.get("expire_date", "")).strip()

                        if expired_date_str:
                            try:
                                expire_date = timezone.datetime.strptime(
                                    expired_date_str, "%Y-%m-%d"
                                ).date()
                            except ValueError:
                                errors.append(
                                    {
                                        "row": row_num,
                                        "name": product_name,
                                        "errors": [
                                            "'expire_date' must be in YYYY-MM-DD format."
                                        ],
                                    }
                                )
                                continue

                        # Upsert variant: match by SKU (if given) else by name+item
                        variant = None
                        if sku:
                            variant = ItemVariant.objects.filter(sku=sku).first()
                            if variant and variant.item != item:
                                errors.append(
                                    {
                                        "row": row_num,
                                        "name": product_name,
                                        "errors": [
                                            f"SKU '{sku}' belongs to a different product."
                                        ],
                                    }
                                )
                                continue
                        if variant is None:
                            variant = ItemVariant.objects.filter(
                                item=item, name=variant_name
                            ).first()

                        # Quantity is populated through a supplied batch below,
                        # not set directly on the variant.
                        if variant:
                            variant.sku = sku
                            variant.save(update_fields=["sku"])
                        else:
                            variant = ItemVariant.objects.create(
                                item=item,
                                name=variant_name,
                                quantity=0,
                                sku=sku,
                                is_default=is_first_variant,
                            )
                            is_first_variant = False

                        # Record stock as a supplied batch under the import
                        # supply. The signal bumps variant.quantity and supply
                        # totals. A SuppliedItem requires a selling price, so
                        # when none is given we set the quantity on the variant.
                        if quantity > 0:
                            if import_supply is not None and selling_price is not None:
                                SuppliedItem.objects.create(
                                    supply=import_supply,
                                    item=item,
                                    variant=variant,
                                    quantity=quantity,
                                    initial_quantity=quantity,
                                    selling_price=selling_price,
                                    purchase_price=None,
                                    batch_number=batch_number,
                                    product_number=(
                                        sku or f"{item.name} — {variant_name}"
                                    )[:255],
                                    business=business,
                                    expire_date=(
                                        expire_date if expired_date_str else None
                                    ),
                                )
                            else:
                                variant.quantity = quantity
                                variant.save(update_fields=["quantity"])

                        created_variants.append(
                            {
                                "row": row_num,
                                "product": product_name,
                                "variant": variant_name,
                            }
                        )

            except Exception as exc:
                errors.append(
                    {"row": first_row_num, "name": product_name, "errors": [str(exc)]}
                )

        return Response(
            {
                "products_created": len(created_items),
                "variants_processed": len(created_variants),
                "error_count": len(errors),
                "supply_label": import_supply.label if import_supply else None,
                "products": created_items,
                "variants": created_variants,
                "errors": errors,
            },
            status=status.HTTP_207_MULTI_STATUS if errors else status.HTTP_201_CREATED,
        )

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    @action(
        detail=False,
        methods=["get"],
        url_path="export",
        permission_classes=[IsAuthenticated, BranchLevelPermission],
    )
    def export(self, request, *args, **kwargs):
        """
        Export all products for the current branch as CSV or Excel.

        One row per variant. Importing the exported file reproduces the same data.

        Query params:
          export_format : 'csv' | 'xlsx'  (default: xlsx)

        Do not use the query name ``format`` — DRF reserves it for content
        negotiation (``?format=json`` / ``api``); ``?format=xlsx`` is handled
        before this view and returns 404 when no xlsx renderer exists.
        """
        fmt = request.query_params.get("export_format", "xlsx").lower()

        items = (
            self.get_queryset()
            .select_related("group")
            .prefetch_related(
                "variants",
                # Latest batch first so supplied_items[0] is the most recent
                # supply — used to source the variant's selling price below.
                Prefetch(
                    "variants__supplied_items",
                    queryset=SuppliedItem.objects.order_by("-created_at"),
                ),
            )
            .order_by("name")
        )

        def variant_selling_price(variant):
            """Selling price from the latest supply batch."""
            latest = next(iter(variant.supplied_items.all()), None)
            if latest and latest.selling_price:
                return str(latest.selling_price)
            return ""

        # Build flat rows — one per variant
        data_rows = []
        for item in items:
            group_value = item.group.name if item.group else ""
            # Sort in Python to preserve the prefetched supplied_items.
            variants = sorted(item.variants.all(), key=lambda v: v.name)
            if not variants:
                # Emit one placeholder row so the product is not silently lost
                data_rows.append(
                    [
                        item.name,
                        item.description or "",
                        item.inventory_unit,
                        "",  # variant_name (will default to product name on re-import)
                        "",  # selling_price — user must fill in
                        "",  # sku
                        "",  # quantity — no variants, no stock data
                        group_value,
                    ]
                )
            else:
                for variant in variants:
                    data_rows.append(
                        [
                            item.name,
                            item.description or "",
                            item.inventory_unit,
                            "" if variant.name == item.name else variant.name,
                            variant_selling_price(variant),
                            variant.sku or "",
                            variant.quantity,
                            group_value,
                        ]
                    )

        if fmt == "csv":
            response = HttpResponse(content_type="text/csv; charset=utf-8")
            response["Content-Disposition"] = (
                'attachment; filename="products_export.csv"'
            )
            response.write("\ufeff")  # UTF-8 BOM for Excel compatibility
            writer = csv.writer(response)
            writer.writerow(BULK_IMPORT_COLUMNS)
            writer.writerows(data_rows)
            return response

        # Default: xlsx
        wb = openpyxl.Workbook()
        ws, fill, font, border = self._styled_wb_sheet(wb, "Products")
        self._write_xlsx_headers(ws, BULK_IMPORT_COLUMNS, fill, font, border)

        row_font = openpyxl.styles.Font(size=10)

        # Colour alternate product blocks to make grouping visible
        palette = ["FFFFFF", "EBF3FB"]
        current_product = None
        colour_idx = 0
        for row_idx, row in enumerate(data_rows, start=2):
            if row[0] != current_product:
                current_product = row[0]
                colour_idx = (colour_idx + 1) % 2
            row_fill = openpyxl.styles.PatternFill("solid", fgColor=palette[colour_idx])
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = row_font
                cell.border = border
                cell.fill = row_fill

        return self._build_xlsx_response(wb, "products_export.xlsx")


class SupplyViewset(
    ListModelMixin,
    CreateModelMixin,
    RetrieveModelMixin,
    UpdateModelMixin,
    GenericViewSet,
):
    serializer_class = SupplySerializer
    permission_classes = [IsAuthenticated, BranchLevelPermission]
    queryset = Supply.objects.all()
    filterset_class = SupplyFilter

    def get_queryset(self):
        queryset = filter_queryset_by_branch(self.queryset, self.request, "supply")
        return queryset.order_by("-updated_at")

    def get_serializer_class(self):
        if self.action in ("retrieve", "update", "partial_update"):
            return SupplyDetailsSerializer
        return self.serializer_class

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        # Always respond with the full detail view (depth=1 nested objects)
        return Response(
            SupplyDetailsSerializer(
                instance, context=self.get_serializer_context()
            ).data
        )

    @action(detail=True, methods=["post"], url_path="settle_debt")
    @idempotent
    def settle_debt(self, request, *args, **kwargs):
        """
        Settle an outstanding supply debt by recording an actual PURCHASE payment.

        Body:
            payment_method  (UUID, required) – the BusinessPaymentMethod used to pay
            amount          (decimal, optional) – overrides supply.total_cost if provided
        """
        from finances.models import BusinessPaymentMethod, Transaction

        supply = self.get_object()
        supply_ref = f"supply:{supply.id}"

        # Must have an existing DEBT transaction for this supply.
        if not Transaction.objects.filter(
            category=supply_ref, type=Transaction.TransactionType.DEBT
        ).exists():
            return Response(
                {"detail": "No outstanding debt found for this supply."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Prevent double-settlement.
        settled_ref = f"{supply_ref}:paid"
        if Transaction.objects.filter(category=settled_ref).exists():
            return Response(
                {"detail": "This supply debt has already been settled."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        pm_id = request.data.get("payment_method")
        if not pm_id:
            return Response(
                {"detail": "payment_method is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            payment_method = BusinessPaymentMethod.objects.get(pk=pm_id)
        except BusinessPaymentMethod.DoesNotExist:
            return Response(
                {"detail": "Payment method not found."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        raw_amount = request.data.get("amount")
        try:
            amount = float(raw_amount) if raw_amount else float(supply.total_cost)
        except (ValueError, TypeError):
            return Response(
                {"detail": "Invalid amount."}, status=status.HTTP_400_BAD_REQUEST
            )

        tx = Transaction.objects.create(
            type=Transaction.TransactionType.PURCHASE,
            total_paid_amount=round(amount, 2),
            payment_method=payment_method,
            business=supply.business,
            branch=supply.branch,
            category=settled_ref,
        )

        from finances.serializers import TransactionSerializer

        return Response(TransactionSerializer(tx).data, status=status.HTTP_201_CREATED)


class SupplierViewset(ListModelMixin, CreateModelMixin, GenericViewSet):
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated, BranchLevelPermission]
    queryset = Supplier.objects.all()
    filterset_class = SupplierFilter

    def get_queryset(self):
        # Supplier has no branch FK (shared across a business), but access is
        # still gated by whether the user has a branch-level view perm on at
        # least one branch in the targeted business.
        if not accessible_branches(self.request, "supplier").exists():
            return self.queryset.none()
        business = self.request.business
        if not business:
            return self.queryset.none()
        return self.queryset.filter(business=business)

    def perform_create(self, serializer):
        # Supplier is shared across the business; bind it to the active business
        # from the request context rather than trusting the client payload.
        serializer.save(business=self.request.business)


class SupplyItemViewset(
    ListModelMixin,
    CreateModelMixin,
    RetrieveModelMixin,
    UpdateModelMixin,
    DestroyModelMixin,
    GenericViewSet,
):
    """Dedicated CRUD endpoints for SuppliedItem (supply batches).

    Access is gated on the parent Supply's branch permissions:
      - list / retrieve  → ``can_view_supply_branch``
      - create / update / delete → ``can_change_supply_branch``

    Quantity changes on update and delete are automatically synced to the
    owning ItemVariant so that total stock always stays consistent.
    """

    serializer_class = SuppliedItemSerializer
    permission_classes = [IsAuthenticated]
    queryset = SuppliedItem.objects.all()
    filterset_class = SuppliedItemFilter

    def get_serializer_class(self):
        if self.action in ("update", "partial_update"):
            return SuppliedItemUpdateSerializer
        return SuppliedItemSerializer

    def get_queryset(self):
        queryset = self.queryset.select_related("supply__branch", "variant", "item")
        branches = accessible_branches(self.request, "supply")
        queryset = queryset.filter(supply__branch__in=branches)
        return queryset

    # ------------------------------------------------------------------
    # Permission helpers
    # ------------------------------------------------------------------

    def _resolve_supply_branch_from_request(self):
        """Resolve branch from the ``supply`` field in request body/params (for create)."""
        supply_id = self.request.data.get("supply") or self.request.query_params.get(
            "supply_id"
        )
        if not supply_id:
            return None
        supply = Supply.objects.filter(id=supply_id).select_related("branch").first()
        return supply.branch if supply else None

    def _check_change_perm(self, branch):
        return self.request.user.has_perm(
            biz_perm("supply", "change", "branch"), branch
        )

    # ------------------------------------------------------------------
    # Actions
    # ------------------------------------------------------------------

    def create(self, request, *args, **kwargs):
        branch = self._resolve_supply_branch_from_request()
        if branch and not self._check_change_perm(branch):
            return Response(
                {"detail": "You do not have permission to add items to this supply."},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().create(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()

        branch = instance.supply.branch
        if not self._check_change_perm(branch):
            return Response(
                {"detail": "You do not have permission to update this supplied item."},
                status=status.HTTP_403_FORBIDDEN,
            )

        old_quantity = instance.quantity
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)

        new_quantity = serializer.validated_data.get("quantity", old_quantity)
        delta = new_quantity - old_quantity

        with transaction.atomic():
            serializer.save()
            if delta != 0:
                variant = instance.variant
                variant.quantity = max(0, variant.quantity + delta)
                variant.save(update_fields=["quantity", "updated_at"])

        return Response(SuppliedItemSerializer(instance).data)

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()

        branch = instance.supply.branch
        if not self._check_change_perm(branch):
            return Response(
                {"detail": "You do not have permission to delete this supplied item."},
                status=status.HTTP_403_FORBIDDEN,
            )

        instance.delete()

        return Response(status=status.HTTP_204_NO_CONTENT)


class PricingViewset(
    CreateModelMixin, DestroyModelMixin, UpdateModelMixin, GenericViewSet
):
    queryset = ItemVariant.objects.all()
    serializer_class = PricingSerializer
    permission_classes = [IsAuthenticated, BranchLevelPermission]

    def get_queryset(self):
        return Pricing.objects.all()


class GroupViewset(ModelViewSet):
    queryset = Group.objects.all()
    serializer_class = GroupSerializer
    permission_classes = [IsAuthenticated, BusinessLevelPermission]
    filterset_class = GroupFilter

    def get_queryset(self):
        queryset = self.queryset
        business = self.request.business
        if not business:
            return queryset.none()
        if self.request.user.has_perm("can_view_group_business", business):
            return queryset.filter(business=business)
        return queryset.none()


_VARIANT_LIST_SCHEMA = extend_schema(
    summary="List item variants",
    description=(
        "Returns a paginated list of item variants. Supports filtering by branch, "
        "business, item, name, SKU, selling price range, and inventory status."
    ),
    parameters=[
        OpenApiParameter(
            name="low_stock",
            type=OpenApiTypes.BOOL,
            location=OpenApiParameter.QUERY,
            description=(
                "Filter by low-stock status. Pass `true` to return only variants "
                "whose total supplied quantity (sum of all SuppliedItem.quantity) "
                "is at or below the parent item's `notify_below` threshold, or "
                "variants with no supplied items at all. Pass `false` to return "
                "only well-stocked variants."
            ),
            required=False,
        ),
        OpenApiParameter(
            name="expiring",
            type=OpenApiTypes.BOOL,
            location=OpenApiParameter.QUERY,
            description=(
                "Filter by expiry status. Pass `true` to return only variants "
                "that have at least one SuppliedItem with a non-null `expire_date` "
                "and remaining `quantity > 0`. Pass `false` to exclude such variants."
            ),
            required=False,
        ),
    ],
)


class ItemVariantViewset(ModelViewSet):
    queryset = ItemVariant.objects.all()
    serializer_class = ItemVariantSerializer
    permission_classes = [IsAuthenticated, BranchLevelPermission]
    filterset_class = ItemVariantFilter

    @_VARIANT_LIST_SCHEMA
    def list(self, request, *args, **kwargs):
        return super().list(request, *args, **kwargs)

    def get_serializer_class(self):
        if self.action == "retrieve":
            return ItemVariantReadSerializer
        if self.action == "list":
            return ItemVariantReadSerializer
        return self.serializer_class

    def get_queryset(self):
        queryset = self.queryset
        item = self.request.query_params.get("item")
        if item:
            queryset = queryset.filter(item=item)

        queryset = filter_queryset_by_branch(
            queryset, self.request, "itemvariant", branch_field="item__branch"
        )
        # ItemVariantReadSerializer nests properties/pricings/supplied_items and
        # uses depth=2 over item -> (group/business/branch/categories); pull them
        # all in so list/retrieve don't issue per-row queries.
        return (
            queryset.select_related("item__group", "item__business", "item__branch")
            .prefetch_related(
                "supplied_items__supply",
                "properties",
                "pricings",
                "item__categories",
            )
            .order_by(Lower("name"), "id")
        )

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)

        return Response(
            ItemVariantReadSerializer(serializer.instance).data,
            status=status.HTTP_201_CREATED,
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        self.perform_update(serializer)
        return Response(
            ItemVariantReadSerializer(serializer.instance).data,
            status=status.HTTP_200_OK,
        )


class PropertyViewset(
    CreateModelMixin, DestroyModelMixin, UpdateModelMixin, GenericViewSet
):
    queryset = ItemVariant.objects.all()
    serializer_class = PropertySerializer
    permission_classes = [IsAuthenticated, BranchLevelPermission]

    def get_queryset(self):
        return Property.objects.all()


class InventoryMovementViewSet(ModelViewSet):
    """ViewSet for managing inventory movements between branches"""

    queryset = InventoryMovement.objects.all()
    permission_classes = [IsAuthenticated, BranchLevelPermission]

    def get_serializer_class(self):
        if self.action == "create":
            return InventoryMovementCreateSerializer
        return InventoryMovementSerializer

    def get_queryset(self):
        queryset = self.queryset.select_related(
            "from_branch", "to_branch", "business", "requested_by"
        ).prefetch_related("movement_items__supplied_item__item")

        status_filter = self.request.query_params.get("status")

        # Branches where the caller can ``view`` inventory movements. Any
        # movement whose source OR destination is in that set is visible.
        branches = accessible_branches(self.request, "inventorymovement")
        if not branches.exists():
            return queryset.none()

        queryset = queryset.filter(
            Q(from_branch__in=branches) | Q(to_branch__in=branches)
        )

        if status_filter:
            queryset = queryset.filter(status=status_filter)

        return queryset

    @action(detail=True, methods=["post"])
    @idempotent
    def approve(self, request, pk=None):
        """Approve a pending inventory movement"""
        movement = self.get_object()

        if movement.status != "pending":
            return Response(
                {"error": "Only pending movements can be approved"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            movement.status = "approved"
            movement.approved_by = request.user
            movement.approved_at = timezone.now()
            movement.save()

        serializer = self.get_serializer(movement)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    @idempotent
    def ship(self, request, pk=None):
        """Mark movement as shipped and reduce inventory from source"""
        movement = self.get_object()

        if movement.status != "approved":
            return Response(
                {"error": "Only approved movements can be shipped"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        shipped_items = request.data.get("items", [])

        with transaction.atomic():
            for item_data in shipped_items:
                movement_item = movement.movement_items.get(
                    id=item_data["movement_item_id"]
                )
                quantity_shipped = item_data["quantity_shipped"]

                if quantity_shipped > movement_item.quantity_requested:
                    return Response(
                        {
                            "error": f"Shipped quantity cannot exceed requested quantity for {movement_item.supplied_item.item.name}"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Update movement item
                movement_item.quantity_shipped = quantity_shipped
                movement_item.save()

                # Reduce inventory from source
                movement_item.supplied_item.quantity -= quantity_shipped
                movement_item.supplied_item.save()

            movement.status = "shipped"
            movement.shipped_by = request.user
            movement.shipped_at = timezone.now()
            movement.save()

        serializer = self.get_serializer(movement)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    @idempotent
    def receive(self, request, pk=None):
        """Mark movement as received and add inventory to destination"""
        movement = self.get_object()

        if movement.status != "shipped":
            return Response(
                {"error": "Only shipped movements can be received"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        received_items = request.data.get("items", [])

        with transaction.atomic():
            for item_data in received_items:
                movement_item = movement.movement_items.get(
                    id=item_data["movement_item_id"]
                )
                quantity_received = item_data["quantity_received"]

                if quantity_received > movement_item.quantity_shipped:
                    return Response(
                        {
                            "error": f"Received quantity cannot exceed shipped quantity for {movement_item.supplied_item.item.name}"
                        },
                        status=status.HTTP_400_BAD_REQUEST,
                    )

                # Update movement item
                movement_item.quantity_received = quantity_received
                movement_item.save()

                # Add inventory to destination
                # First, try to find existing supply in destination branch
                source_supply = movement_item.supplied_item.supply
                source_label = source_supply.label
                if not source_label:
                    # Group by source supply so all items go to same destination supply
                    source_label = f"supply-{source_supply.id}"
                destination_supply, created = Supply.objects.get_or_create(
                    branch=movement.to_branch,
                    label=source_label,
                )

                # Check if same item already exists in destination supply
                existing_supplied_item = SuppliedItem.objects.filter(
                    supply=destination_supply,
                    item=movement_item.supplied_item.item,
                    batch_number=movement_item.supplied_item.batch_number,
                    product_number=movement_item.supplied_item.product_number,
                ).first()

                if existing_supplied_item:
                    # Add to existing stock
                    existing_supplied_item.quantity += quantity_received
                    existing_supplied_item.save()
                else:
                    # Create new supplied item in destination
                    SuppliedItem.objects.create(
                        supply=destination_supply,
                        item=movement_item.supplied_item.item,
                        quantity=quantity_received,
                        selling_price=movement_item.supplied_item.selling_price,
                        purchase_price=movement_item.supplied_item.purchase_price,
                        batch_number=movement_item.supplied_item.batch_number,
                        product_number=f"{movement_item.supplied_item.product_number}-T{movement.id}",  # Avoid duplicate product numbers
                        expire_date=movement_item.supplied_item.expire_date,
                        man_date=movement_item.supplied_item.man_date,
                        business=movement.business,
                        variant=movement_item.variant,
                    )

            movement.status = "received"
            movement.received_by = request.user
            movement.received_at = timezone.now()
            movement.save()

        serializer = self.get_serializer(movement)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    @idempotent
    def cancel(self, request, pk=None):
        """Cancel a movement (only if pending or approved)"""
        movement = self.get_object()

        if movement.status not in ["pending", "approved"]:
            return Response(
                {"error": "Only pending or approved movements can be cancelled"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        movement.status = "cancelled"
        movement.save()

        serializer = self.get_serializer(movement)
        return Response(serializer.data)


class InventoryMovementItemViewSet(ModelViewSet):
    """ViewSet for individual movement items"""

    queryset = InventoryMovementItem.objects.all()
    serializer_class = InventoryMovementItemSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = self.queryset.select_related("movement", "supplied_item__item")

        movement_id = self.request.query_params.get("movement_id")
        if movement_id:
            queryset = queryset.filter(movement_id=movement_id)

        return queryset


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def inventory_summary(request):
    """
    GET /inventories/summary/

    Returns inventory summary for a business or branch:
      - total_products: number of active items
      - stock_in_hand:  total units across all active items
      - low_stock_count: items whose quantity is at or below their notify_below threshold
    """
    business_id = request.query_params.get("business") or request.query_params.get(
        "business_id"
    )
    branch_id = request.query_params.get("branch") or request.query_params.get(
        "branch_id"
    )

    business = None
    branch = None

    if business_id:
        if not is_valid_uuid(business_id):
            raise ValidationError({"detail": "Invalid business ID format"})
        try:
            business = Business.objects.get(id=business_id)
        except Business.DoesNotExist:
            raise ValidationError({"detail": "Business not found"})
    elif hasattr(request, "business") and request.business:
        business = request.business

    if branch_id:
        if not is_valid_uuid(branch_id):
            raise ValidationError({"detail": "Invalid branch ID format"})
        try:
            branch = Branch.objects.get(id=branch_id)
            if business and branch.business != business:
                raise ValidationError(
                    {"detail": "Branch does not belong to the specified business"}
                )
            if not business:
                business = branch.business
        except Branch.DoesNotExist:
            raise ValidationError({"detail": "Branch not found"})
    elif hasattr(request, "branch") and request.branch:
        branch = request.branch
        if not business:
            business = branch.business

    if not business:
        raise ValidationError(
            {
                "detail": (
                    "Business is required. "
                    "Provide 'business' or 'business_id' query parameter"
                )
            }
        )

    if not request.user.has_perm(
        biz_perm("item", "view", "branch"), branch or business
    ):
        return Response(
            {"detail": "You do not have permission to view inventory summary."},
            status=status.HTTP_403_FORBIDDEN,
        )

    items_qs = Item.objects.filter(business=business, is_active=True)
    if branch:
        items_qs = items_qs.filter(branch=branch)

    total_products = items_qs.count()
    stock_in_hand = (
        SuppliedItem.objects.filter(item__in=items_qs).aggregate(total=Sum("quantity"))[
            "total"
        ]
        or 0
    )
    low_stock_count = (
        items_qs.annotate(total_stock=Sum("variants__supplied_items__quantity"))
        .filter(Q(total_stock__lte=F("notify_below")) | Q(total_stock__isnull=True))
        .count()
    )

    return Response(
        {
            "total_products": total_products,
            "stock_in_hand": stock_in_hand,
            "low_stock_count": low_stock_count,
        }
    )
